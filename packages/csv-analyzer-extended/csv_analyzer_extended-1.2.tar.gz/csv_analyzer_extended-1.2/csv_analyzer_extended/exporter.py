import csv
import json
import os
import multiprocessing as mp
import sqlite3  # Für die SQLite-Integration
import openpyxl  # Für den Excel-Export
import pyarrow as pa  # Für den Parquet-Export
import pyarrow.parquet as pq  # Für den Parquet-Export
from fpdf import FPDF

class CSVExporter:
    def __init__(self, data, header):
        self.data = data
        self.header = header

    def export_to_csv(self, file_path):
        with open(file_path, 'w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(self.header)
            writer.writerows(self.data)

    def export_to_json(self, file_path):
        json_data = [dict(zip(self.header, row)) for row in self.data]
        with open(file_path, 'w', encoding='utf-8') as file:
            json.dump(json_data, file, indent=4)

    def export_to_excel(self, file_path):
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(self.header)
        for row in self.data:
            sheet.append(row)
        workbook.save(file_path)

    def export_to_html(self, file_path):
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write('<table>\n')
            file.write('  <tr>' + ''.join([f'<th>{col}</th>' for col in self.header]) + '</tr>\n')
            for row in self.data:
                file.write('  <tr>' + ''.join([f'<td>{cell}</td>' for cell in row]) + '</tr>\n')
            file.write('</table>')

    def export_to_parquet(self, file_path):
        table = pa.Table.from_arrays([pa.array([row[i] for row in self.data]) for i in range(len(self.header))],
                                     names=self.header)
        pq.write_table(table, file_path)

    def export_to_markdown(self, file_path):
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write('| ' + ' | '.join(self.header) + ' |\n')
            file.write('|' + '---|' * len(self.header) + '\n')
            for row in self.data:
                file.write('| ' + ' | '.join(row) + ' |\n')

    def export_to_latex(self, file_path):
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write('\\begin{tabular}{' + ' | '.join(['l'] * len(self.header)) + '}\n')
            file.write('  ' + ' & '.join(self.header) + ' \\\\\n')
            file.write('  \\hline\n')
            for row in self.data:
                file.write('  ' + ' & '.join(row) + ' \\\\\n')
            file.write('\\end{tabular}\n')
            
    def export_to_pdf(self, file_path):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Kopfzeile
        pdf.set_font("Arial", 'B', 12)
        for col in self.header:
            pdf.cell(40, 10, col, 1, 0, 'C')
        pdf.ln()

        # Datenzeilen
        pdf.set_font("Arial", size=12)
        for row in self.data:
            for item in row:
                pdf.cell(40, 10, item, 1, 0, 'C')
            pdf.ln()

        pdf.output(file_path)

    def export_to_sql(self, table_name, cursor):
        """Exportiert die Daten in eine SQL-Datenbank."""
        columns = ', '.join([f'"{col}"' for col in self.header])
        placeholders = ', '.join(['?'] * len(self.header))
        insert_query = f'INSERT INTO {table_name} ({columns}) VALUES ({placeholders})'
        for row in self.data:
            cursor.execute(insert_query, row)    
                
    def export_to_xml(self, file_path):
        """Exportiert die Daten in eine XML-Datei."""
        with open(file_path, 'w', encoding='utf-8') as file:
            file.write('<?xml version="1.0" encoding="UTF-8"?>\n')
            file.write('<data>\n')
            for row in self.data:
                file.write('  <row>\n')
                for tag, value in zip(self.header, row):
                    file.write(f'    <{tag}>{value}</{tag}>\n')
                file.write('  </row>\n')
            file.write('</data>\n')
