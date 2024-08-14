
import numpy as np
import pandas as pd
import plotly.graph_objects as go
from pyvis.network import Network
from scipy.stats import norm
from sklearn.model_selection import train_test_split, GridSearchCV
from sklearn.linear_model import LogisticRegression
from sklearn.metrics import classification_report, confusion_matrix
from sklearn.kernel_approximation import RBFSampler

from openpyxl import Workbook
from openpyxl.utils import get_column_letter as letter
from openpyxl.styles import Font, Color, PatternFill, NamedStyle
from openpyxl.formatting.rule import ColorScale, FormatObject, Rule



"""
Class for evaluating cluster leakage in predictions data.

Parameters:
- predictions: DataFrame containing prediction data.
- id_col: Column name for example IDs.
- target_col: Column name for target values.
- pred_col: Column name for predicted values.
- prob_col: Column name for probability values.

Methods:
- get_leakage_counts(detection_thresh: float = 0.05) -> np.ndarray: Get leakage count matrix.
- get_support() -> np.ndarray: Get support of each target.
- get_leakage(detection_thresh: float = 0.05) -> np.ndarray: Get leakage matrix.
- get_leakage_dictionary(detection_thresh: float = 0.05, leakage_thresh: float = 0.02, printPhrases = True) -> dict: Create dictionary of leakages above a threshold.
- save_leakage_graph(detection_thresh: float = 0.05, leakage_thresh: float = 0.02, filename: str = "cluster_leakage_graph.html"): Save interactive leakage graph to HTML.
- get_total_leakage(detection_thresh: float = 0.05) -> float: Get total leakage percentage.
- hypothesis_test_total_leakage(significance_level = 0.05): Perform hypothesis testing on total leakage.
- save_leakage_report(detection_thresh: float = 0.05, leakage_thresh: float = 0.02, significance_level: float = 0.05, filename: str = "cluster_leakage_report.xlsx"): Write a leakage report to an excel file.
"""
class ClusterEvaluator:


    """
    Initialize the ClusterEvaluator object with prediction data after preprocessing.

    Parameters:
    - predictions: DataFrame containing prediction data.
    - id_col: Column name for example IDs (default is "ID").
    - target_col: Column name for target values (default is "TARGET").
    - pred_col: Column name for predicted values (default is "PREDICTION").
    - prob_col: Column name for probability values (default is "PROBABILITY").
    """
    def __init__(self,  
                    predictions: pd.DataFrame, 
                    id_col: str = "ID", 
                    target_col: str = "TARGET",
                    pred_col: str = "PREDICTION",
                    prob_col: str = "PROBABILITY"):

        self.predictions: pd.DataFrame = self.__prep_data__(predictions, id_col, target_col, pred_col, prob_col)
        self.support: np.ndarray = None
        self.num_clusters: int = predictions["TARGET"].nunique()
        self.count = predictions.shape[0]
        self.targets_arr = self.predictions[["TARGET"]].to_numpy().reshape(-1).astype(int)
        self.off_prob_arr = self.predictions.drop(["ID", "TARGET", "PREDICTION"], axis=1).to_numpy()
        self.off_prob_arr[np.arange(self.count), self.targets_arr] = 0


    """
    Prepare the prediction data by ensuring correct column types and renaming columns as needed.

    Parameters:
    - predictions: DataFrame containing prediction data.
    - id_col: Column name for example IDs (default is "ID").
    - target_col: Column name for target values (default is "TARGET").
    - pred_col: Column name for predicted values (default is "PREDICTION").
    - prob_col: Column name for probability values (default is "PROBABILITY").

    Returns:
    DataFrame: Processed prediction data with corrected column types and renamed columns.
    """
    def __prep_data__(self,
                        predictions: pd.DataFrame, 
                        id_col: str = "ID", 
                        target_col: str = "TARGET",
                        pred_col: str = "PREDICTION",
                        prob_col: str = "PROBABILITY"):
        
        actual_types = predictions.dtypes.to_dict()
        expected_types = {
            id_col: "object",
            target_col: "int32",
            pred_col: "int32",
            prob_col: "object"
        }

        for col_name, col_type in expected_types.items():
            if col_name not in actual_types:
                raise ValueError(f"The given dataframe is missing the column '{col_name}'!")
            elif actual_types[col_name] != col_type:
                raise ValueError(f"The column {col_name} has incorrect type {actual_types[col_name]}! Expected: {col_type}")

        predictions = predictions[[id_col, target_col, pred_col, prob_col]].rename(columns={
            id_col: "ID",
            target_col: "TARGET",
            pred_col: "PREDICTION",
            prob_col: "PROBABILITY"
        })

        prob_df = pd.DataFrame(predictions["PROBABILITY"].to_list(), index=predictions.index)
        old_names = prob_df.columns
        prob_df = prob_df.rename(columns={
            old_names[i]: str(i)
            for i in range(len(old_names))
        })
        predictions = predictions.drop("PROBABILITY", axis=1)
        predictions = pd.concat([predictions, prob_df], axis=1)

        return predictions


    """
    Calculate the count of leakage instances between clusters based on a given detection threshold.

    Parameters:
    - detection_thresh: The threshold off-target probability for detecting leakage (default is 0.05).

    Returns:
    np.ndarray: A matrix of leakage counts between clusters.
    """
    def get_leakage_counts(self, detection_thresh: float = 0.05) -> np.ndarray:

        agg_dict = {
            str(i): (lambda x: (x >= float(detection_thresh)).sum()) 
            for i in range(self.num_clusters)
        }
        leakage_counts = self.predictions.groupby("TARGET").agg(agg_dict).reset_index()
        leakage_counts = leakage_counts.sort_values(by="TARGET")

        return leakage_counts.drop("TARGET", axis=1).to_numpy()
    

    """
    Get the support of each target cluster.

    Returns:
    np.ndarray: Array containing the support of each target, ordered from 0 upwards.
    """
    def get_support(self) -> np.ndarray:

        if self.support is not None:
            return self.support
        
        support_df = self.predictions.groupby("TARGET").size().reset_index(name='count').sort_values(by="TARGET")
        self.support = support_df[["count"]].to_numpy().reshape(-1)

        return self.support
    

    """
    Calculate the leakage percentage between clusters based on a given detection threshold.

    Parameters:
    - detection_thresh: The threshold off-target probability for detecting leakage (default is 0.05).

    Returns:
    np.ndarray: Array containing the leakage percentages between clusters.
    """
    def get_leakage(self, detection_thresh: float = 0.05) -> np.ndarray:

        counts = self.get_leakage_counts(detection_thresh)
        support = self.get_support().reshape(-1, 1)

        return counts / support


    """
    Get a dictionary of cluster leakages above a specified threshold.

    Parameters:
    - detection_thresh: The threshold off-target probability for detecting leakage (default is 0.05).
    - leakage_thresh: The threshold for leakage percentage to be included in the dictionary (default is 0.02).
    - printPhrases: Boolean indicating whether to print readable phrases (default is True).

    Returns:
    dict: A dictionary where keys are tuples representing cluster pairs leaking, and values are tuples containing leakage percentage, leakage count, support, and a descriptive phrase.
    """
    def get_leakage_dictionary(self, 
                                 detection_thresh: float = 0.05, 
                                 leakage_thresh: float = 0.02,
                                 printPhrases = True) -> dict:
        
        # Set diagonal to 0 so that it isn't counted. (Cluster does not "influence" itself).
        leakage_matrix = self.get_leakage(detection_thresh)
        np.fill_diagonal(leakage_matrix, 0)

        leakage_counts = self.get_leakage_counts(detection_thresh)
        support_arr = self.get_support()

        leakage_dict = {}

        for i in range(self.num_clusters):
            for j in range(self.num_clusters):
                if leakage_matrix[i][j] >= leakage_thresh:
                    phrase = f"Cluster {i} leaks into cluster {j} by {leakage_matrix[i][j]:.2%}  ({leakage_counts[i][j]} / {support_arr[i]})"
                    leakage_dict[(i, j)] = (leakage_matrix[i][j], leakage_counts[i][j], support_arr[i], phrase)

        # Optionally print readable phrases.
        if printPhrases:
            influence_dict_sorted = sorted(leakage_dict.items(), key=lambda item: item[1][0], reverse=True)
            for item in influence_dict_sorted:
                _, _, _, phrase = item[1]
                print(phrase)

        return leakage_dict
    

    """
    Save an interactive leakage graph to an HTML file.

    Parameters:
    - detection_thresh: The threshold for off-target probability to detect leakage (default is 0.05).
    - leakage_thresh: The threshold for leakage percentage to be included in the graph (default is 0.02).
    - filename: The name of the HTML file to save the graph (default is "cluster_leakage_graph.html").
    """
    def save_leakage_graph(self,
                               detection_thresh: float = 0.05,
                               leakage_thresh: float = 0.02,
                               filename: str = "cluster_leakage_graph.html"):

        if not filename.endswith(".html"):
            filename = filename + ".html"

        leakage_matrix = self.get_leakage(detection_thresh)
        np.fill_diagonal(leakage_matrix, 0) # No self-edges in the graph!

        # Create a directed graph
        net = Network(notebook=True, directed=True, height="750px", width="100%", bgcolor="#222222", font_color="black")

        # Add nodes.
        for i in range(self.num_clusters):
            net.add_node(i, label=str(i), shape="circle")

        # Function to determine color brightness based on weight
        def get_color(weight):
            weight = weight / 0.05
            base_color = "255,255,255"  # White in RGB
            intensity = max(0, min(255, int(255 * weight)))  # Adjust intensity based on weight
            return f"rgba({base_color},{intensity})"

        # Add edges with weights, adjusting for bidirectional edges to avoid overlap
        for i in range(self.num_clusters):
            for j in range(self.num_clusters):
                if leakage_matrix[i][j] >= leakage_thresh:
                    color = get_color(leakage_matrix[i][j])
                    edge_label = f"{leakage_matrix[i][j]:.2%}"
                    if leakage_matrix[j][i] >= leakage_thresh:  # Check for bidirectional edge
                        # For bidirectional edges, use "smooth" type to avoid overlap
                        net.add_edge(i, j, title=edge_label, label=edge_label, width=3, 
                                     smooth={'type': 'curvedCW', 'roundness': 0.2}, arrowStrikethrough=False, color=color)
                    else:
                        # For unidirectional edges, use straight lines
                        net.add_edge(i, j, title=edge_label, label=edge_label, width=3, arrowStrikethrough=False, color=color)

        # Set options for the graph to make it more interactive
        net.set_options(
        """
        var options = {
        "nodes": {
            "shape": "circle",
            "size": 50,
            "font": {
            "size": 18,
            "face": "arial",
            "strokeWidth": 0,
            "align": "center",
            "color": "black"
            }
        },
        "edges": {
            "arrows": {
            "to": {
                "enabled": true,
                "scaleFactor": 1
            }
            },
            "color": {
            "inherit": false
            },
            "smooth": {
            "enabled": true,
            "type": "dynamic"
            },
            "font": {
            "size": 14,
            "align": "top",
            "strokeWidth": 0,
            "color": "white"
            },
            "width": 2
        },
        "physics": {
            "forceAtlas2Based": {
            "gravitationalConstant": -100,
            "centralGravity": 0.01,
            "springLength": 100,
            "springConstant": 0.08
            },
            "maxVelocity": 50,
            "solver": "forceAtlas2Based",
            "timestep": 0.35,
            "stabilization": {
            "enabled": true,
            "iterations": 1000,
            "updateInterval": 25,
            "onlyDynamicEdges": false,
            "fit": true
            }
        }
        }
        """
        )

        net.save_graph(filename)


    """
    Calculate the total leakage percentage based on a given detection threshold.

    Parameters:
    - detection_thresh: The threshold off-target probability for detecting leakage (default is 0.05).

    Returns:
    float: The total leakage percentage.
    """
    def get_total_leakage(self, detection_thresh: float = 0.05) -> float:

        max_off_prob = np.max(self.off_prob_arr, axis=1)
        num_leaks = np.sum(max_off_prob >= detection_thresh)

        return num_leaks / self.count
    

    """
    Perform hypothesis testing on the total leakage at different detection thresholds and comparison values.
    
    Parameters:
    - significance_level: The significance level for the hypothesis test (default is 0.05).

    Returns:
    Tuple: A tuple containing a plotly figure, cell labels, cell colors, x-axis values, and y-axis values.
    """
    def hypothesis_test_total_leakage(self, significance_level = 0.05):

        comparison_values = np.linspace(0.05, 0.25, num=9)
        detection_thresholds = np.linspace(0.05, 0.25, num=9)

        xlen = len(detection_thresholds)
        ylen = len(comparison_values)

        total_leakage_stats = np.zeros(xlen)
        for i in range(xlen):
            total_leakage_stats[i] = self.get_total_leakage(detection_thresholds[i])

        # The null hypothesis is "the total leakage is equal to X"
        # The alternative is "the total leakage is more than X"
        # Performing the grid of hypothesis tests:

        alpha = significance_level

        n = self.count
        p_0 = comparison_values.reshape(-1, 1)
        p_hat = total_leakage_stats.reshape(1, -1)

        numerator = np.repeat(p_hat, ylen, axis=0) - np.repeat(p_0, xlen, axis=1)
        denominator = np.sqrt(p_0 * (1 - p_0) / n)
        denominator = np.repeat(denominator, xlen, axis=1)
        z_stats = numerator / denominator

        p_values = 1 - norm.cdf(z_stats)
        decisions = p_values < alpha

        # Format decisions in a table.

        cell_labels = np.round(np.repeat(p_hat, ylen, axis=0), decimals=4)
        cell_colors = decisions.astype(int)
        x_axis = np.round(detection_thresholds, decimals=4)
        y_axis = np.round(comparison_values, decimals=4)

        fig = go.Figure(data=go.Heatmap(
            z=cell_colors,
            text=cell_labels,
            texttemplate="%{text}",
            colorscale=[[0, "green"], [1, "red"]],
            showscale=False  # This line removes the colorbar
        ))

        fig.update_layout(
            title="Hypothesis Testing for Cluster Leakage",
            xaxis=dict(title="Detection Threshold", tickvals=list(range(len(x_axis))), ticktext=x_axis),
            yaxis=dict(title="Comparison", tickvals=list(range(len(y_axis))), ticktext=y_axis)
        )

        return (fig, cell_labels, cell_colors, x_axis, y_axis)
    
    
    """
    Save a leakage report to an Excel file with detailed statistics and conditional formatting.

    Parameters:
    - detection_thresh: The threshold for off-target probability to detect leakage (default is 0.05).
    - leakage_thresh: The threshold for leakage percentage to be colored in the report (default is 0.02).
    - significance_level: The significance level for hypothesis testing (default is 0.05).
    - filename: The name of the Excel file to save the report (default is "cluster_leakage_report.xlsx").
    """
    def save_leakage_report(self, 
                        detection_thresh: float = 0.05, 
                        leakage_thresh: float = 0.02,
                        significance_level: float = 0.05,
                        filename: str = "cluster_leakage_report.xlsx"):

        if not filename.endswith(".xlsx"):
            filename = filename + ".xlsx"

        col_names = ["ID", "TARGET", "PREDICTION"]
        predictions_arr = self.predictions.to_numpy()
        leakage_arr = self.get_leakage(detection_thresh)
        _, cell_labels, cell_colors, x_axis, y_axis = self.hypothesis_test_total_leakage(significance_level)

        wb = Workbook()

        # ---------------------------------------------------------------------   Write to Sheet 1.
        sheet1 = wb.active
        sheet1.title = "Predictions and Targets"

        # Header row
        sheet1.freeze_panes = "A2" # Freeze header row.
        header = col_names + [f"CLUSTER_{i}" for i in range(self.num_clusters)]
        sheet1.append(header)

        header_style = NamedStyle(name="header")
        header_style.font = Font(bold=True)
        header_style.fill = PatternFill(patternType="solid", fgColor="9ec0ff")

        for col in sheet1.iter_cols(min_row=1, max_row=1):
            col[0].style = header_style

        # Add the prediction data.
        for i in range(predictions_arr.shape[0]):
            sheet1.append(list(predictions_arr[i]))

        last_row = 1 + predictions_arr.shape[0]
        last_col = letter(sheet1.max_column)
        last_cell_address = f"{last_col}{last_row}"

        percent_style = NamedStyle(name="percent")
        percent_style.number_format = "0.00%"
        cells = sheet1[f"D2:{last_cell_address}"]
        for row in cells:
            for cell in row:
                cell.style = percent_style

        # Conditional color gradient formatting for the probability outputs.
        first = FormatObject(type='num', val=0)
        last = FormatObject(type='num', val=1)
        colors = [Color('FFFFFF'), Color('02bd56')]
        color_scale = ColorScale(cfvo=[first, last], color=colors)
        color_rule = Rule(type='colorScale', colorScale=color_scale)
        sheet1.conditional_formatting.add(f"D2:{last_cell_address}", color_rule)
    
        # -----------------------------------------------------------------------------  Write to sheet 2.
        sheet2 = wb.create_sheet("Cluster Leakage")

        # Cluster Title Row
        cluster_names = [f"CLUSTER_{i}" for i in range(self.num_clusters)]
        i = 0
        for col in sheet2.iter_cols(min_row=1, max_row=1, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = cluster_names[i]
            col[0].style = header_style
            i = i + 1

        # Statistic Titles
        cells = sheet2["C2":"C6"]
        titles = ["SUPPORT", "COUNT > 5%", "LEAKY RECALL", "RECALL", "LEAKAGE"]
        i = 0
        for row in cells:
            for cell in row:
                cell.value = titles[i]
                cell.style = header_style
                i = i + 1

        # Fill in support.
        i = 0
        for col in sheet2.iter_cols(min_row=2, max_row=2, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"=COUNTIF('Predictions and Targets'!B:B, \"={i}\")"
            i = i + 1

        # Fill in Count > 5%.
        i = 0
        for col in sheet2.iter_cols(min_row=3, max_row=3, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"=COUNTIF('Predictions and Targets'!{letter(4+i)}:{letter(4+i)}, \">0.05\")"
            i = i + 1

        # Fill in Leaky Recall.
        i = 0
        for col in sheet2.iter_cols(min_row=4, max_row=4, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"={letter(4+i)}3 / {letter(4+i)}2"
            col[0].style = percent_style
            i = i + 1

        # Fill in Recall.
        i = 0
        for col in sheet2.iter_cols(min_row=5, max_row=5, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"=COUNTIFS('Predictions and Targets'!B:B, \"={i}\", 'Predictions and Targets'!C:C, \"={i}\") / {letter(4+i)}2"
            col[0].style = percent_style
            i = i + 1

        # Fill in Leakage percentage.
        i = 0
        for col in sheet2.iter_cols(min_row=6, max_row=6, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = f"={letter(4+i)}4 - {letter(4+i)}5"
            col[0].style = percent_style
            i = i + 1

        # Create axes for leakage matrix.
        axes_label_fill = PatternFill(patternType="solid", fgColor="ed9fd4")
        for col in sheet2.iter_cols(min_row=11, max_row=11, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].fill = axes_label_fill
        for row in sheet2.iter_rows(min_row=13, max_row=(13 + self.num_clusters - 1), min_col=2, max_col=2):
            row[0].fill = axes_label_fill
        sheet2["F11"] = "LEAKS TO CLUSTER"
        sheet2["B15"] = "TARGET CLUSTER"
        i = 0
        for col in sheet2.iter_cols(min_row=12, max_row=12, min_col=4, max_col=(4 + self.num_clusters - 1)):
            col[0].value = cluster_names[i]
            col[0].style = header_style
            i = i + 1
        i = 0
        for row in sheet2.iter_rows(min_row=13, max_row=(13 + self.num_clusters - 1), min_col=3, max_col=3):
            row[0].value = cluster_names[i]
            row[0].style = header_style
            i = i + 1

        # Add the inter-cluster leakages.
        matrix_corner = f"{letter(4 + self.num_clusters - 1)}{13 + self.num_clusters - 1}"
        cells = sheet2["D13":matrix_corner]
        for i in range(len(cells)):
            for j in range(len(cells[0])):
                cells[i][j].value = leakage_arr[i][j]
                cells[i][j].style = percent_style

        # Delete the diagonal of the matrix.
        black_fill = PatternFill(patternType="solid", fgColor="000000")
        for i in range(self.num_clusters):
            address = f"{letter(4 + i)}{13 + i}"
            sheet2[address] = ""
            sheet2[address].fill = black_fill

        # Gradient conditional formatting
        first = FormatObject(type='num', val=leakage_thresh)
        last = FormatObject(type='max')
        colors = [Color('FFFFFF'), Color('02bd56')]
        color_scale = ColorScale(cfvo=[first, last], color=colors)
        color_rule = Rule(type='colorScale', colorScale=color_scale)
        sheet2.conditional_formatting.add(f"D13:{matrix_corner}", color_rule)

        # --------------------------------------------------------------------------------- Write to Sheet 3.
        sheet3 = wb.create_sheet("Total Leakage Tests")

        # Create axes for leakage matrix.
        for col in sheet3.iter_cols(min_row=2, max_row=2, min_col=4, max_col=(4 + 9 - 1)):
            col[0].fill = axes_label_fill
        for row in sheet3.iter_rows(min_row=4, max_row=(4 + 9 - 1), min_col=2, max_col=2):
            row[0].fill = axes_label_fill
        sheet3["F2"] = "DETECTION THRESHOLD"
        sheet3["B6"] = "COMPARISON"
        i = 0
        for col in sheet3.iter_cols(min_row=3, max_row=3, min_col=4, max_col=(4 + 9 - 1)):
            col[0].value = x_axis[i]
            col[0].style = header_style
            i = i + 1
        i = 0
        for row in sheet3.iter_rows(min_row=4, max_row=(4 + 9 - 1), min_col=3, max_col=3):
            row[0].value = y_axis[i]
            row[0].style = header_style
            i = i + 1

        # Fill in the matrix with the total leakage and hypotheses.
        matrix_corner = f"{letter(4 + 9 - 1)}{4 + 9 - 1}"
        cells = sheet3["D4":matrix_corner]
        NO_fill = PatternFill(patternType="solid", fgColor="76e393")
        YES_fill = PatternFill(patternType="solid", fgColor="e37676")
        for i in range(len(cells)):
            for j in range(len(cells[0])):
                cells[i][j].value = cell_labels[i][j]
                cells[i][j].style = percent_style
                cells[i][j].fill = NO_fill if cell_colors[i][j] == 0 else YES_fill

        # Small key for coloring.
        sheet3["D15"] = "REJECT"
        sheet3["D15"].fill = YES_fill
        sheet3["D16"] = "FAIL TO REJECT"
        sheet3["D16"].fill = NO_fill
        sheet3["E15"] = "Statistically significant evidence that cell value (total leakage) is larger than comparison value."
        sheet3["E16"] = "No evidence that cell value (total leakage) is larger than comparison value."

        # -------------------------------------------------------------------- Cleanup and Save.

        # Iterate over all columns and adjust their widths
        for ws_name in wb.sheetnames:
            for column in wb[ws_name].columns:
                column_letter = column[0].column_letter
                wb[ws_name].column_dimensions[column_letter].width = 15

        wb.save(filename)



"""
Class for predicting clusters based on given data.

Parameters:
- clustering_data: DataFrame containing the data for clustering.
- id_col: Column name for IDs.
- target_col: Column name for target values.
- nonlinear_boundary: Boolean indicating whether to use a nonlinear boundary (default True).

Methods:
- evaluate_model_train(): Evaluate the model on the train set.
- evaluate_model_test(): Evaluate the model on the test set.
- evaluate_model(): Perform full evaluation on both train and test sets.
- get_train_predictions(): Get formatted DataFrame of predictions on the train set.
- get_test_predictions(): Get formatted DataFrame of predictions on the test set.
"""
class ClusterPredictor:


    """
    Initialize the ClusterPredictor class with the provided clustering data and parameters.

    Parameters:
    - clustering_data: DataFrame containing the data for clustering.
    - id_col: Column name for IDs (default is "ID").
    - target_col: Column name for target values (default is "TARGET").
    - nonlinear_boundary: Boolean indicating whether to use a nonlinear boundary (default is True).

    The method processes the data, trains a model, and stores model predictions for train and test sets.
    """
    def __init__(self,
                 clustering_data: pd.DataFrame, 
                 id_col: str = "ID", 
                 target_col: str = "TARGET",
                 nonlinear_boundary: bool = True):

        # Keep track of data.
        data_objects = self.__process_data__(clustering_data, id_col, target_col)
        data_labels = ["train_data", "test_data", "train_x", "train_y", "test_x", "test_y"]
        self.data = {
            data_labels[i]: data_objects[i]
            for i in range(6)
        }
        

        # Train a model on the data.
        self.nonlinear_boundary = nonlinear_boundary
        self.model = self.__train_model__()

        # Keep track of model predictions on the internal train and test sets.
        self.data["train_y_pred"] = self.model.predict(self.data["train_x"])
        self.data["train_y_pred_proba"] = self.model.predict_proba(self.data["train_x"])
        self.data["test_y_pred"] = self.model.predict(self.data["test_x"])
        self.data["test_y_pred_proba"] = self.model.predict_proba(self.data["test_x"])
    

    """
    Process the clustering data by checking and adjusting column types, renaming columns, and splitting the data into train and test sets.

    Parameters:
    - clustering_data: DataFrame containing the data for clustering.
    - id_col: Column name for IDs (default is "ID").
    - target_col: Column name for target values (default is "TARGET").

    Returns:
    - Tuple containing train_data, test_data, train_x, train_y, test_x, test_y.
    """
    def __process_data__(self,
                         clustering_data: pd.DataFrame, 
                         id_col: str = "ID", 
                         target_col: str = "TARGET"):
        
        actual_types = clustering_data.dtypes.to_dict()
        expected_types = {
            id_col: "object",
            target_col: "int32"
        }

        for col_name, col_type in expected_types.items():
            if col_name not in actual_types:
                raise ValueError(f"The given dataframe is missing the column '{col_name}'!")
            elif actual_types[col_name] != col_type:
                raise ValueError(f"The column {col_name} has incorrect type {actual_types[col_name]}! Expected: {col_type}")

        clustering_data = clustering_data.rename(columns={
            id_col: "ID",
            target_col: "TARGET"
        })

        
        # Split the data into train and test sets
        train_data, test_data = train_test_split(clustering_data, test_size=0.2, random_state=42)

        feature_cols = clustering_data.columns.to_list()
        feature_cols.remove("ID")
        feature_cols.remove("TARGET")

        train_x = train_data[feature_cols].to_numpy()
        train_y = train_data[["TARGET"]].to_numpy().reshape(-1)
        test_x = test_data[feature_cols].to_numpy()
        test_y = test_data[["TARGET"]].to_numpy().reshape(-1)

        return train_data, test_data, train_x, train_y, test_x, test_y
    
    
    """
    Train a logistic regression model with optional nonlinear boundary transformation using RBFSampler.

    Returns:
        The best estimator model after performing grid search cross-validation.
    """
    def __train_model__(self):

        if self.nonlinear_boundary:

            rbf = RBFSampler(gamma='scale', random_state=42)
            self.data["train_x"] = rbf.fit_transform(self.data["train_x"])
            self.data["test_x"] = rbf.transform(self.data["test_x"])

        log_reg = LogisticRegression(random_state=42)

        param_grid = {
            'C': np.logspace(-4, 0, 5)  # 5 values from 10^-4 to 10^0
        }

        crossval = GridSearchCV(estimator=log_reg, 
                             param_grid=param_grid,
                             cv=5,
                             scoring='accuracy')
        
        crossval.fit(self.data["train_x"], self.data["train_y"])

        return crossval.best_estimator_
        

    """
    Evaluate the trained model on the train set by generating a classification report and a confusion matrix visualized as a heatmap. The confusion matrix displays the true labels on the y-axis and the predicted labels on the x-axis. The heatmap provides a visual representation of the classification performance on the training data.
    """
    def evaluate_model_train(self):

        y_true = self.data["train_y"]
        y_pred = self.data["train_y_pred"]

        print("\n----------------------- Train Evaluation -----------------------")

        # Generate the classification report
        report = classification_report(y_true, y_pred)
        print("Classification Report:\n", report)

        # Generate the confusion matrix
        conf_matrix = confusion_matrix(y_true, y_pred)
        conf_matrix = np.round(conf_matrix, decimals=4)
        fig = go.Figure(data=go.Heatmap(
            z=conf_matrix,
            text=conf_matrix,
            texttemplate="%{text}"
        ))

        fig.update_layout(
            title="Confusion Matrix (TRAIN)",
            xaxis=dict(title="Prediction", tickvals=list(range(len(self.model.classes_)))),
            yaxis=dict(title="Label", tickvals=list(range(len(self.model.classes_))))
        )

        fig.show()

        print("----------------------------------------------------------------\n")


    """
    Evaluate the trained model on the test set by generating a classification report and a confusion matrix visualized as a heatmap. The confusion matrix displays the true labels on the y-axis and the predicted labels on the x-axis. The heatmap provides a visual representation of the classification performance on the test data.
    """
    def evaluate_model_test(self):

        y_true = self.data["test_y"]
        y_pred = self.data["test_y_pred"]

        print("\n----------------------- Test Evaluation -----------------------")

        # Generate the classification report
        report = classification_report(y_true, y_pred)
        print("Classification Report:\n", report)

        # Generate the confusion matrix
        conf_matrix = confusion_matrix(y_true, y_pred)
        conf_matrix = np.round(conf_matrix, decimals=4)
        fig = go.Figure(data=go.Heatmap(
            z=conf_matrix,
            text=conf_matrix,
            texttemplate="%{text}"
        ))

        fig.update_layout(
            title="Confusion Matrix (TEST)",
            xaxis=dict(title="Prediction", tickvals=list(range(len(self.model.classes_)))),
            yaxis=dict(title="Label", tickvals=list(range(len(self.model.classes_))))
        )

        fig.show()

        print("----------------------------------------------------------------\n")


    """
    Execute the evaluation of the trained model on both the train and test sets.
    """
    def evaluate_model(self):
        self.evaluate_model_train()
        self.evaluate_model_test()


    """
    Generate and return a DataFrame with predictions for the training set.

    Returns:
        pd.DataFrame: DataFrame containing columns 'ID', 'TARGET', 'PREDICTION', and 'PROBABILITY'.
    """
    def get_train_predictions(self):

        output: pd.DataFrame = pd.DataFrame(self.data["train_data"][["ID", "TARGET"]])
        output["PREDICTION"] = pd.Series(self.data["train_y_pred"],
                                         index=output.index)
        output["PROBABILITY"] = pd.Series(self.data["train_y_pred_proba"].tolist(), 
                                          index=output.index)

        return output
    

    """
    Generate and return a DataFrame with predictions for the test set.

    Returns:
        pd.DataFrame: DataFrame containing columns 'ID', 'TARGET', 'PREDICTION', and 'PROBABILITY'.
    """
    def get_test_predictions(self):

        output: pd.DataFrame = pd.DataFrame(self.data["test_data"][["ID", "TARGET"]])
        output["PREDICTION"] = pd.Series(self.data["test_y_pred"], 
                                         index=output.index)
        output["PROBABILITY"] = pd.Series(self.data["test_y_pred_proba"].tolist(), 
                                          index=output.index)

        return output