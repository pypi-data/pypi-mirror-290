from openpyxl.styles import Border, Side, PatternFill, Font, Alignment

# Default styles
TITLE_COLOR = PatternFill(patternType='solid', fgColor='00000080')
CATEGORY_COLOR = PatternFill(patternType='solid', fgColor='0096C8FA')
TABLE_COLOR = PatternFill(patternType='solid', fgColor='00D9D9D9')
CASE_COLOR = PatternFill(patternType='solid', fgColor='00C9C9C9')
PASSED_COLOR = PatternFill(patternType='solid', fgColor='0000AF00')
FAILED_COLOR = PatternFill(patternType='solid', fgColor='00EF0000')
NOT_APPLICABLE_COLOR = PatternFill(patternType='solid', fgColor='00AAAAAA')

BORDERS_ALL = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
BORDERS_LEFT_RIGHT = Border(left=Side(style='thin'), right=Side(style='thin'))
BORDERS_LEFT_RIGHT_TOP = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'))
BORDERS_TOP = Border(top=Side(style='thin'))

DEFAULT_ROW_HEIGHT = 15

def SetTitle(ts, line, text):
    SetSection(ts, line, text, TITLE_COLOR)
    ts[f"B{line}"].font = Font(color='00FFFFFF', bold=True, size=12)
    
def SetCategory(ts, line, text):
    SetSection(ts, line, text, CATEGORY_COLOR)
    
def SetCategoryValue(ts, line, value):
    valueRange = f"B{line}:F{line}"
    ts.merge_cells(valueRange)
    SetBorder(ts, valueRange, BORDERS_ALL)
    if type(value) is list: # If list of string, join with carriage return and adjust row height
        ts.row_dimensions[line].height = len(value) * DEFAULT_ROW_HEIGHT
        value = '\n'.join(value)
    ts[f"B{line}"] = value
    ts[f"B{line}"].alignment = Alignment(wrap_text=True, horizontal = "left", vertical = "top")
    
    
def SetSection(ts, line, text, fill):
    ts[f"B{line}"] = text
    ts[f"B{line}"].fill = fill
    ts.merge_cells(f"B{line}:F{line}")
    ts[f"B{line}"].font = Font(bold=True)

def SetBoldCell(ts, range, text, fill, border):
    SetCell(ts, range, text, fill, border)
    ts[range].font = Font(bold=True)

def SetBoldCell(ts, range, text, fill, border):
    SetCell(ts, range, text, fill, border)
    if type(range) is str:
        ts[range].font = Font(bold=True)
    elif type(range) is list:
        row = range[0]
        column = range[1]
        ts.cell(row=row, column=column).font = Font(bold=True)
    
def SetLongVariable(ts, line, variable, value):
    variableCell = ts[f"B{line}"]
    variableCell.value = variable
    variableCell.border = BORDERS_ALL
    valueRange = f"C{line}:F{line}"
    ts.merge_cells(valueRange)
    SetBorder(ts, valueRange, BORDERS_ALL)
    valueCell = ts[f"C{line}"]
    if type(value) is list: # If list of string, join with carriage return and adjust row height
        ts.row_dimensions[line].height = len(value) * DEFAULT_ROW_HEIGHT
        value = '\n'.join(value)
    valueCell.value = value
    valueCell.border = BORDERS_ALL
    valueCell.alignment = Alignment(wrap_text=True, horizontal = "left", vertical = "top")
    
def SetCell(ts, range, text, fill, border):
    if type(range) is str:
        ts[range] = text
        if fill != None: ts[range].fill = fill
        ts[range].alignment = Alignment(wrap_text=True, horizontal = "left", vertical = "top")
        if border != None: ts[range].border = border
    elif type(range) is list:
        row = range[0]
        column = range[1]
        ts.cell(row=row, column=column).value= text
        if fill != None: ts.cell(row=row, column=column).fill = fill
        ts.cell(row=row, column=column).alignment = Alignment(wrap_text=True, horizontal = "left", vertical = "top")
        if border != None: ts.cell(row=row, column=column).border = border

def SetFont(ts, range, color = "00000000", bold = False , size = 11):
    if type(range) is str:
        ts[range].font = Font(color=color, bold=bold, size=size)
    elif type(range) is list:
        row = range[0]
        column = range[1]
        ts.cell(row=row, column=column).font = Font(color=color, bold=bold, size=size)


def SetBorder(ts, range, border):
    for row in ts[range]:
        for cell in row:
            cell.border = border