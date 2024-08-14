import os
from openpyxl.workbook import Workbook
from openpyxl.comments import Comment
from openpyxl.utils.cell import get_column_letter
from pathlib import Path
from . import Traceability 
from termcolor import colored
from . import Formatters
from enum import Enum
from datetime import datetime

class Sheet:
    global TsWorkbook
    global Ts

    # Attributes
    global Name
    global Reference
    global Version
    global Description
    global StartConditions
    global Logs

    # Sheet states
    class SheetStates(Enum):
        INIT = 1
        CASE = 2
        ACTION = 3
        EXPECTED_RESULT = 4
        RESULT = 5
    global SheetState
    
    # Sheet modes
    class Modes(Enum):
        Specification = 0
        Execution = 1
    global Mode

    class SheetStatistics:
        PassedCount = 0
        FailedCount = 0
        GlobalStatus = None
        PlayedAt = None
        PlayedBy = ""

    global Statistics

    CaseNr = 0
    ActionNr = 0
    ExpectedResultNr = 0
    LineNr = 0

    FIRST_CASE_LINE = 18

    # Looping
    IsLooping = False
    LoopTargetName = None
    LoopNumber = 0
    LoopStartCaseNr = 0
    LoopStartActionNr = 0
    LoopStartLine = None
    LOOP_FIRST_COLUMN_IDX = 10
    LoopColumn = LOOP_FIRST_COLUMN_IDX
    
    # Conditional Execution
    IsCaseExecuted = True

    class Log:
        global Type
        global Filename

        def __init__(self, type, filename):
            Type = type
            Filename = filename
    
    def __init__(self, name):
        """ Create a new test sheet with reference set to <specId>_<name>
        specId has been set when configuring scripts through Scripts.Configure()
        The name is the name of the test sheet (e.g. RT_EXIT for 'Exit routes')"""

        prefix = os.environ["TEST_SPEC_ID"]
        self.Reference = f"{prefix}_{Path(name).stem}"
        mode = os.environ["MODE"]
        print(colored(f"Creating sheet '{self.Reference}' in '{mode}' mode", "blue"))

        self.Version = os.environ["TEST_SPEC_VERSION"]
        self.Description = "<Undefined>"
        self.StartConditions = "<Undefined>"
        self.Logs = []
        
        self.Mode = self.Modes.Specification if os.environ["MODE"] == "Specification" else "Execution"

        self.TsWorkbook = Workbook()
        self.Ts = self.TsWorkbook.active
        
        self.CaseNr = 0
        self.ActionNr = 0
        self.ExpectedResultNr = 0
        self.LineNr = self.FIRST_CASE_LINE

        self.SheetState = self.SheetStates.INIT
        self.Statistics = self.SheetStatistics()
        
        Traceability.Initialize()

    def Case(self, text):
        """ Create a new 'Case' step in the test sheet with specified 'text'"""

        self._AllowSheetState(self.SheetState.CASE)
        
        self.CaseNr +=1
        self.ActionNr = 0

        textRange = f"C{self.LineNr}:H{self.LineNr}"
        Formatters.SetBorder(self.Ts, textRange, Formatters.BORDERS_ALL)
        self.Ts.merge_cells(textRange)
        Formatters.SetCell(self.Ts, f"B{self.LineNr}", chr(64 + self.CaseNr), Formatters.CASE_COLOR, Formatters.BORDERS_ALL)
        Formatters.SetCell(self.Ts, f"C{self.LineNr}", text, Formatters.CASE_COLOR, Formatters.BORDERS_ALL)
        self.LineNr +=1
        
        self.IsCaseExecuted = True
        
    
    def Action(self, text):
        """ Create a new 'Action' step in the test sheet with specified 'text'"""

        self._AllowSheetState(self.SheetState.ACTION)

        self.ActionNr +=1
        self.ExpectedResultNr = 0

        Formatters.SetCell(self.Ts, f"B{self.LineNr}", self.ActionNr, None, Formatters.BORDERS_LEFT_RIGHT_TOP)
        Formatters.SetCell(self.Ts, f"C{self.LineNr}", text, None, Formatters.BORDERS_LEFT_RIGHT_TOP)

    
    def ExpectedResult(self, text, requirements: list = None, parameters:list = None): 
        """ Create a new 'Expected result' associated to the last 'Action' with specified 'text'
        Optionally, you can pass:
         - The 'requirements' argument containing a list of Req ID tags with syntax [xxxx]
         - The 'parameters' argument containing a list of Xpath tags with syntax <Xpath>
        Note: Make sure the syntax of the requirements & parameters is correct otherwise exceptions will be thrown """

        self._AllowSheetState(self.SheetState.EXPECTED_RESULT)
        
        self.ExpectedResultNr += 1
        
        # Format cells below action
        if self.Ts[f"B{self.LineNr}"].value is None:
            self.Ts[f"B{self.LineNr}"].border = Formatters.BORDERS_LEFT_RIGHT
            self.Ts[f"C{self.LineNr}"].border = Formatters.BORDERS_LEFT_RIGHT
        # Fill expected result
        Formatters.SetCell(self.Ts, f"D{self.LineNr}" , text, None, Formatters.BORDERS_LEFT_RIGHT_TOP)

        # Fill covers
        if requirements != None:
            Formatters.SetCell(self.Ts, f"H{self.LineNr}", '\n'.join(requirements), None, Formatters.BORDERS_ALL) 
        
        Formatters.SetBorder(self.Ts, f"D{self.LineNr}:H{self.LineNr}", Formatters.BORDERS_LEFT_RIGHT_TOP)
        self.LineNr +=1

        # Update traceability
        if requirements != None:
            for requirement in requirements:
                Traceability.AddRequirement(requirement, self.Reference, self.LineNr, self.LoopTargetName)

        if parameters != None:
            for parameter in parameters:
                Traceability.AddParameter(parameter, self.Reference, self.LineNr, self.LoopTargetName)
    
    def Result(self, isPassed: bool, comment: str = None, note: str = None) :
        """ Writes the result of the test, if MODE = 'Execution'
        The 'isPassed' argument is mandatory and shall be set to True if test is passed, False if not.
        Optionally, you can pass the 'comment' argument with a string providing details on the test result."""

        # Column Status = E (5)
        # Column Comment = F (6)

        if self.Mode == self.Modes.Specification: return

        self._AllowSheetState(self.SheetState.RESULT)
        
        self.LineNr -=1

        # Statistics
        if self.IsCaseExecuted:
            if isPassed:
                if self.Statistics.GlobalStatus == None: self.Statistics.GlobalStatus = True
                self.Statistics.PassedCount += 1
            else:
                self.Statistics.GlobalStatus = False
                self.Statistics.FailedCount += 1
        
        # Result & Comment
        if not self.IsLooping:
            if self.IsCaseExecuted:
                self._WriteStatus(isPassed, 5)
                Formatters.SetCell(self.Ts, [self.LineNr, 6], comment, None, Formatters.BORDERS_ALL)
                Formatters.SetFont(self.Ts, [self.LineNr, 6], size = 7) 
                if note != None:
                    excelComment = Comment(note, 'SheetCode')
                    self.Ts.cell(row=self.LineNr, column=6).comment = excelComment
            else:
                self._WriteNotApplicable(5)
        else:
            if self.IsCaseExecuted:
                self._WriteStatus(isPassed, self.LoopColumn, comment)
                if note != None:
                    excelComment = Comment(note, 'SheetCode')
                    self.Ts.cell(row=self.LineNr, column=self.LoopColumn).comment = excelComment
                # Line sub-global status
                if self.Ts.cell(row=self.LineNr, column=5).value == None: 
                    self._WriteStatus(isPassed, 5, isLogged=False)
                else:
                    if not isPassed: self._WriteStatus(isPassed, 5, isLogged=False)
            else:
                self._WriteNotApplicable(self.LoopColumn)
                
        self.LineNr +=1

    def _WriteStatus(self, isPassed, column, text = None, isLogged:bool = True):
        if text == None:
            text = "Passed" if isPassed else "Failed"
     
        colorArgb = Formatters.PASSED_COLOR if isPassed else Formatters.FAILED_COLOR
        colorName = 'green' if isPassed else 'red'

        Formatters.SetCell(self.Ts, [self.LineNr, column], text, colorArgb, Formatters.BORDERS_ALL) 
        if isLogged: print(colored(f"Line {self.LineNr}: Case {chr(64 + self.CaseNr)}, Action {self.ActionNr}, Expected Result {self.ExpectedResultNr} - Passed", colorName))
        
        
    def _WriteNotApplicable(self, column):
        Formatters.SetCell(self.Ts, [self.LineNr, column], "N.A.", Formatters.NOT_APPLICABLE_COLOR, Formatters.BORDERS_ALL) 

    def Save(self):
        print(colored(f"Saving sheet '{self.Reference}'", "blue"))

        # Columns width
        self.Ts.column_dimensions['A'].width = 7
        self.Ts.column_dimensions['B'].width = 14
        self.Ts.column_dimensions['C'].width = 56
        self.Ts.column_dimensions['D'].width = 56
        self.Ts.column_dimensions['E'].width = 11
        self.Ts.column_dimensions['F'].width = 33
        self.Ts.column_dimensions['G'].width = 33
        self.Ts.column_dimensions['H'].width = 33
        
        # Content        
        Formatters.SetTitle(self.Ts, 1, self.Name)
        
        Formatters.SetCategory(self.Ts, 3, "Overview")
        
        Formatters.SetLongVariable(self.Ts, 5, "Name", self.Name)
        Formatters.SetLongVariable(self.Ts, 6, "Reference", self.Reference)
        Formatters.SetLongVariable(self.Ts, 7, "Version", self.Version)
        Formatters.SetLongVariable(self.Ts, 8, "Description", self.Description)
        
        Formatters.SetCategory(self.Ts, 11, "Start conditions")

        Formatters.SetCategoryValue(self.Ts, 13, self.StartConditions)
                
        Formatters.SetCategory(self.Ts, 15, "Test execution")
        
        Formatters.SetBoldCell(self.Ts, f"B{self.FIRST_CASE_LINE-1}","Step", Formatters.TABLE_COLOR, Formatters.BORDERS_ALL)
        Formatters.SetBoldCell(self.Ts, f"C{self.FIRST_CASE_LINE-1}","Action", Formatters.TABLE_COLOR, Formatters.BORDERS_ALL)
        Formatters.SetBoldCell(self.Ts, f"D{self.FIRST_CASE_LINE-1}","Expected result",Formatters.TABLE_COLOR, Formatters.BORDERS_ALL)
        Formatters.SetBoldCell(self.Ts, f"E{self.FIRST_CASE_LINE-1}","Status", Formatters.TABLE_COLOR, Formatters.BORDERS_ALL)
        Formatters.SetBoldCell(self.Ts, f"F{self.FIRST_CASE_LINE-1}","Comment", Formatters.TABLE_COLOR, Formatters.BORDERS_ALL)
        Formatters.SetBoldCell(self.Ts, f"G{self.FIRST_CASE_LINE-1}","Test case", Formatters.TABLE_COLOR, Formatters.BORDERS_ALL)
        Formatters.SetBoldCell(self.Ts, f"H{self.FIRST_CASE_LINE-1}","Covers", Formatters.TABLE_COLOR, Formatters.BORDERS_ALL)

        # Hide "AbsoluteNr" (First) and "Test case" column
        self.Ts.column_dimensions["A"].hidden= True
        self.Ts.column_dimensions["G"].hidden= True

        # Auto filters
        filters = self.Ts.auto_filter
        filters.ref = f"B{self.FIRST_CASE_LINE-1}:H{self.FIRST_CASE_LINE-1}"

        # Add top border of last step line
        Formatters.SetBorder(self.Ts, f"B{self.LineNr}:H{self.LineNr}", Formatters.BORDERS_TOP)

        # Add global status
        self.LineNr +=2
        Formatters.SetCategory(self.Ts, self.LineNr, "Global Status")
        if self.Statistics.GlobalStatus == None:
            Formatters.SetLongVariable(self.Ts, self.LineNr + 2, "Status", "")
            Formatters.SetLongVariable(self.Ts, self.LineNr + 3, "Comment", "")
        else:
            self.Statistics.PlayedAt = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
            self.Statistics.PlayedBy = os.getlogin()
            globalStatusText = "Passed" if self.Statistics.GlobalStatus else "Failed"
            Formatters.SetLongVariable(self.Ts, self.LineNr + 2, "Status", globalStatusText)
            Formatters.SetLongVariable(self.Ts, self.LineNr + 3, "Comment", 
                                    f"Number of Passed : {self.Statistics.PassedCount}\n"
                                    f"Number of Failed : {self.Statistics.FailedCount}\n"
                                    f"Played at : {self.Statistics.PlayedAt}\n"
                                    f"Played by : {self.Statistics.PlayedBy}\n")
            globalStatusColor = "green" if self.Statistics.GlobalStatus else "red"
            print(colored(f"[Global Status : {globalStatusText}]", globalStatusColor))

        self.Ts.row_dimensions[self.LineNr + 3].height = 5 * Formatters.DEFAULT_ROW_HEIGHT

        # Check RVT and Parameters coverage (if requested)
        Traceability.CheckRvtCoverage(self.Reference)
        Traceability.CheckParametersCoverage(self.Reference)

        # Save coverage
        Traceability.SaveRvtCoverage(self.Reference)
        Traceability.SaveParametersCoverage(self.Reference)

        # Adjust zoom level
        self.Ts.sheet_view.zoomScale = 85
               
        testSheetsDirectory = os.environ["TEST_SHEETS_PATH"] 
        if not os.path.exists(testSheetsDirectory): os.mkdir(testSheetsDirectory)
        self.TsWorkbook.save(os.path.join(testSheetsDirectory, f"{self.Reference}_{self.Version}.xlsx"))

    def _AllowSheetState(self, newState):
        # Forbidden transitions
        if self.SheetState == self.SheetStates.INIT and newState != self.SheetStates.CASE:
            raise Exception("a 'Case' must be created first.")
        elif self.SheetState == self.SheetStates.CASE and newState == self.SheetStates.CASE:
            raise Exception("a 'Case' must not be empty. You must have called 'Case' twice in a row.")
        elif self.SheetState == self.SheetStates.ACTION and newState == self.SheetStates.ACTION:
            raise Exception("a 'Action' must have at least one 'Expected Result'")
        elif self.SheetState == self.SheetStates.EXPECTED_RESULT and newState != self.SheetStates.RESULT and os.environ["MODE"].lower() == "execution":
            raise Exception("an 'Expected Result' must have a 'Result' in Execution mode")
        elif self.SheetState == self.SheetStates.RESULT and newState == self.SheetStates.RESULT:
            raise Exception("There must be only one 'Result' per 'Expected Result'. You must have called 'Result' twice in a row.")    
        
        if not(newState == self.SheetStates.RESULT and not os.environ["MODE"].lower() == "execution"):
            self.SheetState = newState

    def StartLoop(self, loopTargetName):
        if self.IsLooping == False:
            self.IsLooping = True
            self.LoopNumber = 1
            self.LoopStartCaseNr = self.CaseNr
            self.LoopStartActionNr = self.ActionNr
            self.LoopStartLine = self.LineNr
            self.LoopColumn = self.LOOP_FIRST_COLUMN_IDX
        else:
            self.LoopNumber +=1
            self.CaseNr = self.LoopStartCaseNr
            self.ActionNr = self.LoopStartActionNr
            self.LineNr = self.LoopStartLine
            self.LoopColumn += 1

        Formatters.SetBoldCell(self.Ts, [self.FIRST_CASE_LINE-1, self.LoopColumn], loopTargetName, Formatters.TABLE_COLOR, Formatters.BORDERS_ALL)
        self.Ts.column_dimensions[get_column_letter(self.LoopColumn)].width = 15

        self.LoopTargetName = loopTargetName

        print(colored(f"[Loop #{self.LoopNumber} on '{loopTargetName}']", "blue"))

    def StopLoop(self):
        self.IsLooping = False
        self.LoopTargetName = None
        self.LoopNumber = 0




        
